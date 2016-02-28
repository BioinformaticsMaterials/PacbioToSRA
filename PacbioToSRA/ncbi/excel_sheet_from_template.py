import os
import shutil

from openpyxl import load_workbook


# TODO: logging


class ExcelSheetFromTemplate(object):

    NCBI_SRA_SUBMISSION_FILE = os.path.join(
        os.path.dirname(os.path.realpath(__file__)),
        'templates',
        'SRA_submission.xlsx'
    )
    FILES_WORKSHEET_NAME = 'Sheet1'
    SR_DATA_WORKSHEET_NAME = 'SRA_data'

    def __init__(self, filename):
        if os.path.exists(filename):
            raise Exception("File already exists: {}".format(filename))

        self._output_file = filename
        self._copy_template(self._output_file)

        self._output_wb = load_workbook(filename=self._output_file)
        self._files_ws = self._output_wb.get_sheet_by_name(self.FILES_WORKSHEET_NAME)
        self._sr_data_ws = self._output_wb.get_sheet_by_name(self.SR_DATA_WORKSHEET_NAME)

    def _copy_template(self, output_file):
        """Copies the template to a new file.

        :param  output_file:    Output file
        :type   output_file:    string
        """
        shutil.copy(self.NCBI_SRA_SUBMISSION_FILE, output_file)

    def _save_output_workbook(self):
        """ Saves output file.
        """
        self._output_wb.save(self._output_file)

    def write_to_files_worksheet(self, rows):
        """Writes the rows to the excel sheet that lists all the files."

        Columns headers:
            Library_ID
            Run_ID
            Filename
            md5sum

        :param  rows:   Rows to write
        :type   rows:   list
        """
        self._do_write_to_worksheet(self._files_ws, rows)

    def write_to_sr_data_worksheet(self, rows):
        """Writes the rows to the excel sheet that has the data on the sample run."

        Columns headers:
            bioproject_accession
            biosample_accession
            sample_name
            library_ID
            title/short description
            library_strategy (click for details)
            library_source (click for details)
            library_selection (click for details)
            library_layout
            platform (click for details)
            instrument_model
            design_description
            reference_genome_assembly (or accession)
            alignment_software
            forward_read_length
            reverse_read_length
            filetype
            filename
            MD5_checksum
            filetype
            filename
            MD5_checksum

        :param  rows:   Rows to write
        :type   rows:   list
        """

        self._do_write_to_worksheet(self._sr_data_ws, rows)

    def _do_write_to_worksheet(self, ws, rows):
        """Writes rows to the worksheet.

        :param  ws:     Worksheet
        :type   ws:     worksheet
        :param  rows:   Rows to write
        :type   rows    list
        """

        # assumes that each worksheet has a header
        row_index = 2           # excel worksheets start at 1. skip header

        for row in rows:
            col_index = 1

            for entry in row:
                ws.cell(row=row_index, column=col_index).value = entry
                col_index += 1

            row_index += 1

        self._save_output_workbook()

